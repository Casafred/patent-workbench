"""
测试权利要求导出增强功能

测试新增的独立权利要求汇总工作表和引用关系JSON格式
"""

import os
import sys
import json
import pandas as pd
from openpyxl import load_workbook

# 添加项目路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from patent_claims_processor.services import ProcessingService, ExportService
from patent_claims_processor.models import ProcessedClaims, ClaimInfo


def test_export_enhancement():
    """测试导出增强功能"""
    
    print("=" * 60)
    print("测试权利要求导出增强功能")
    print("=" * 60)
    print()
    
    # 创建测试数据
    print("1. 创建测试数据...")
    test_claims = [
        # 第一组：来自同一原始单元格的权利要求
        ClaimInfo(
            claim_number=1,
            claim_type='independent',
            claim_text='一种智能手机，包括处理器和存储器。',
            language='中文',
            referenced_claims=[],
            original_text='1. 一种智能手机，包括处理器和存储器。\n2. 根据权利要求1所述的智能手机，其特征在于还包括显示屏。\n3. 根据权利要求1或2所述的智能手机，其特征在于还包括摄像头。',
            confidence_score=0.95
        ),
        ClaimInfo(
            claim_number=2,
            claim_type='dependent',
            claim_text='根据权利要求1所述的智能手机，其特征在于还包括显示屏。',
            language='中文',
            referenced_claims=[1],
            original_text='1. 一种智能手机，包括处理器和存储器。\n2. 根据权利要求1所述的智能手机，其特征在于还包括显示屏。\n3. 根据权利要求1或2所述的智能手机，其特征在于还包括摄像头。',
            confidence_score=0.92
        ),
        ClaimInfo(
            claim_number=3,
            claim_type='dependent',
            claim_text='根据权利要求1或2所述的智能手机，其特征在于还包括摄像头。',
            language='中文',
            referenced_claims=[1, 2],
            original_text='1. 一种智能手机，包括处理器和存储器。\n2. 根据权利要求1所述的智能手机，其特征在于还包括显示屏。\n3. 根据权利要求1或2所述的智能手机，其特征在于还包括摄像头。',
            confidence_score=0.90
        ),
        # 第二组：来自另一个原始单元格的权利要求
        ClaimInfo(
            claim_number=4,
            claim_type='independent',
            claim_text='一种手机壳，包括保护层和装饰层。',
            language='中文',
            referenced_claims=[],
            original_text='4. 一种手机壳，包括保护层和装饰层。\n5. 根据权利要求4所述的手机壳，其特征在于保护层为硅胶材质。',
            confidence_score=0.93
        ),
        ClaimInfo(
            claim_number=5,
            claim_type='dependent',
            claim_text='根据权利要求4所述的手机壳，其特征在于保护层为硅胶材质。',
            language='中文',
            referenced_claims=[4],
            original_text='4. 一种手机壳，包括保护层和装饰层。\n5. 根据权利要求4所述的手机壳，其特征在于保护层为硅胶材质。',
            confidence_score=0.91
        ),
    ]
    
    processed_claims = ProcessedClaims(
        total_cells_processed=5,
        total_claims_extracted=5,
        language_distribution={'中文': 5},
        independent_claims_count=2,
        dependent_claims_count=3,
        processing_errors=[],
        claims_data=test_claims
    )
    
    print(f"   创建了 {len(test_claims)} 个测试权利要求")
    print(f"   独立权利要求: {processed_claims.independent_claims_count}")
    print(f"   从属权利要求: {processed_claims.dependent_claims_count}")
    print()
    
    # 测试导出功能
    print("2. 测试Excel导出...")
    export_service = ExportService()
    
    try:
        output_buffer, filename = export_service.generate_excel_bytesio(processed_claims)
        print(f"   ✓ Excel文件生成成功: {filename}")
        print(f"   ✓ 文件大小: {len(output_buffer.getvalue())} bytes")
        
        # 保存到临时文件以便检查
        temp_file = os.path.join('output', 'test_export_enhancement.xlsx')
        os.makedirs('output', exist_ok=True)
        
        with open(temp_file, 'wb') as f:
            f.write(output_buffer.getvalue())
        
        print(f"   ✓ 测试文件已保存: {temp_file}")
        print()
        
        # 验证Excel文件内容
        print("3. 验证Excel文件内容...")
        wb = load_workbook(temp_file)
        
        # 检查工作表
        sheet_names = wb.sheetnames
        print(f"   工作表列表: {sheet_names}")
        
        expected_sheets = ['独立权利要求汇总', '权利要求详情', '处理统计']
        for sheet_name in expected_sheets:
            if sheet_name in sheet_names:
                print(f"   ✓ 工作表 '{sheet_name}' 存在")
            else:
                print(f"   ✗ 工作表 '{sheet_name}' 不存在")
        print()
        
        # 检查汇总工作表
        print("4. 检查独立权利要求汇总工作表...")
        if '独立权利要求汇总' in sheet_names:
            summary_sheet = wb['独立权利要求汇总']
            
            # 读取数据
            df = pd.read_excel(temp_file, sheet_name='独立权利要求汇总')
            print(f"   汇总表行数: {len(df)}")
            print(f"   汇总表列: {list(df.columns)}")
            print()
            
            # 显示汇总数据
            print("   汇总数据预览:")
            for idx, row in df.iterrows():
                print(f"   行 {idx + 1}:")
                print(f"      原数据行号: {row['原数据行号']}")
                print(f"      独立权利要求序号: {row['独立权利要求序号']}")
                print(f"      独立权利要求文本: {row['独立权利要求文本'][:50]}...")
                print(f"      引用关系: {row['从属权利要求引用关系']}")
                
                # 验证JSON格式
                try:
                    ref_map = json.loads(row['从属权利要求引用关系'])
                    print(f"      ✓ JSON格式有效，包含 {len(ref_map)} 个引用关系")
                except json.JSONDecodeError as e:
                    print(f"      ✗ JSON格式无效: {e}")
                print()
        
        print("=" * 60)
        print("测试完成！")
        print("=" * 60)
        
        return True
        
    except Exception as e:
        print(f"   ✗ 测试失败: {e}")
        import traceback
        print(traceback.format_exc())
        return False


if __name__ == '__main__':
    success = test_export_enhancement()
    sys.exit(0 if success else 1)
